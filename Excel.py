from openpyxl import *
import datetime
import time
class Excel:

  def __init__(self,file):
    self.file=file
    self.book = load_workbook(filename = file)
    #TODO check that this file actually exists
  def writeToFile(self,package,primary):
    """
    Package format
    ["SN's",[primary],[]
    """
    sn=package[0]
    current_meas=package[1]
    voltage_meas=package[2]
    led_shorting=package[3]
    pressure_sensor=package[4]
    #need to check if sheet exists
    sheet_name = str(sn)+"&"+str(sn)
    if sheet_name in self.book.sheetnames:
        target=self.book[sheet_name]
    else:
        try:
          target = self.book.copy_worksheet(self.book['Blank'])
        except:
          raise Exception("Invalid Excell document, must include the 'Blank' worksheet to duplicate")
        target.title = sheet_name
    if primary:
        cell_positions={"sn":"b2","datetime":"c2","current_col":2,"voltage_col":2,"led_shorting":"C17","pressure_sensor":"D22"}
    else:
        cell_positions={"sn":"b3","datetime":"c3","current_col":4,"voltage_col":3,"led_shorting":"C18","pressure_sensor":None}
    #check if we already have values written
    if target[cell_positions["sn"]].value:
        print("This board already appears to have been tested")
        inp=int(input("1) Overwrite Data\n2) Create new worksheet\n>>"))
        if inp == 1:
          pass
        elif inp == 2:
          target = self.book.copy_worksheet(self.book['Blank'])
          target.title = sheet_name+"_"+str(int(time.time()))
        else:
          raise Exception("Invalid selection")
    #Write SN
    target[cell_positions["sn"]]=sn
    #Write Test datetime
    target[cell_positions["datetime"]]=str(datetime.datetime.now())
    #write current stuff
    r = 8
    for n in current_meas:
        target.cell(row=r, column=cell_positions["current_col"]).value = n
        r += 1
    #write voltage_meas
    r = 27
    for n in voltage_meas:
        target.cell(row=r, column=cell_positions["voltage_col"]).value = n
        r += 1
    #write shorting
    target[cell_positions["led_shorting"]]=led_shorting
    #write pressure sensor
    if cell_positions["pressure_sensor"] != None:
        target[cell_positions["pressure_sensor"]]=pressure_sensor

        
  def saveFile(self):
    
    self.book.save(self.file)
